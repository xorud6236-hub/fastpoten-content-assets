# -*- coding: utf-8 -*-
"""CA-6 자체 테스트 — 팩트 자리(스키마) + 팩트 적재(추가 전용·검수 보존).

★ 실제 창고(data/content_assets.sqlite3)에는 절대 쓰지 않는다 — 전부 임시 DB.
사용: python tests/test_rulebook_facts.py   (표준 unittest, 추가 의존성 없음)
"""
import os
import sqlite3
import sys
import tempfile
import unittest

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "src"))
import db  # noqa: E402
import load_rulebook  # noqa: E402
import masking  # noqa: E402


def _fake_workbook(individual_rows, common_rows=None, ind_name_header="상품/키워드명"):
    """②시트와 같은 모양의 임시 엑셀(메모리) — 새 항목 추가·변경 감지 테스트용.

    ind_name_header: 개별 팩트의 항목명 머리글. 엑셀에서 머리글을 고친 상황을 흉내낼 때 바꾼다.
    """
    common_rows = common_rows if common_rows is not None else [COMMON_ROW_1]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = load_rulebook.FACT_SHEET
    ws.append(["SECTION A — 공통 팩트 (카테고리 단위 공통 적용)"])
    ws.append(["No.", "구분", "카테고리", "응시/취득 요건", "필요 학점", "예상 소요 기간",
               "기간 단축 방법", "자주 묻는 질문 TOP3", "주의사항 / 흔한 오해",
               "주의메모 (시점/예외)"])
    for r in common_rows:
        ws.append(list(r))
    ws.append([])
    ws.append(["SECTION B — 개별 팩트 (키워드/상품별 세부 정보)"])
    ws.append(["No.", "구분", "연결 카테고리", ind_name_header, "핵심 팩트",
               "학력별 경로 요약", "글 작성 시 강조포인트", "주의메모", "사용 우선순위", "비고"])
    for r in individual_rows:
        ws.append(list(r))
    ws.append(["▼ 아래부터 개별 팩트를 추가해주세요 ▼"])
    return wb


IND_ROW_1 = (1, "개별", "테스트카테고리", "테스트상품", "실습 120시간",
             "고졸: 학점은행제", "조건까지 함께", "주의메모 원본", "높음", "비고")
# 공통 팩트는 항목명이 곧 '카테고리' 칸(3번째) — 이름 바뀜 테스트에서 이 칸만 고친다
COMMON_ROW_1 = (1, "공통", "테스트카테고리", "학사학위 필요", "140학점", "2년",
                "전적대 인정", "Q1 몇 시간? → 120시간", "흔한 오해 없음", "2020.1.1 이전은 120시간")


def _full_workbook_file(tmp, name, individual_rows):
    """run() 검증용 — ①카테고리·④금지어 시트까지 갖춘 임시 엑셀 '파일'(run은 경로를 받는다).

    individual_rows=[] 로 부르면 개별 팩트 0건 = 깨진 엑셀(parse_fact_rows가 멈춘다).
    """
    wb = _fake_workbook(individual_rows)
    ws = wb.create_sheet(load_rulebook.CATEGORY_SHEET)
    ws.append(["No.", "상위 카테고리", "카테고리명", "키워드 예시", "학점은행제 링크",
               "고유 키워드 수", "총 발행 빈도", "활성"])
    ws.append([1, "복지", "테스트카테고리", "예시 키워드", "https://example.test", 3, 10, "Y"])
    ws = wb.create_sheet(load_rulebook.BANNED_SHEET)
    ws.append(["금지어 목록"])
    ws.append(["No.", "금지어", "사유", "대체 표현"])
    ws.append([1, "무조건", "단정 표현", "대체로"])
    path = os.path.join(tmp, name)
    wb.save(path)
    return path


# 실제 staff 테이블의 '모양'만 흉내낸 가공 이름 — 2~3자 한글 이름에 이름 아닌 값이 섞여 있다.
# (실제 staff엔 1글자·점 붙은 값·업체명이 섞여 있다. 그 모양이 마스킹 오탐의 씨앗이라 재현한다.)
# ★ 실명 반입 금지(불변 1): 이 저장소는 공개다. 실제 직원 이름을 여기 넣지 말 것 —
#   테스트를 만들 때 실제 명단을 그대로 베끼기 쉬우니 특히 주의(2026-07-15에 실제로 그랬다).
# 마스킹은 이 목록을 그대로 쓴다(load_staff_names) → 팩트 적재가 이 경로를 실제로 밟는다.
STAFF_SAMPLE = ["가상인", "나철수", "다영희", "라민호", "철수", "영희",
                "9", "가", ".마가가", "가상교육원"]

# 팩트 시트에 실제로 들어 있는 모양의 문장 — 마스킹이 절대 건드리면 안 되는 것들
FACT_SENTENCES = [
    "2020.1.1 이전 입학자는 실습 120시간, 이후는 160시간입니다.",
    "고졸: 학점은행제로 140학점 이수 후 응시 가능",
    "예상 소요 기간 1년 6개월 (전적대 학점 인정 시 단축)",
    "Q1. 실습은 몇 시간인가요? → 2020.1.1 이후 160시간",
    "필요 학점 17학점 · 사회복지 현장실습 필수",
]


def _insert_staff(conn):
    for n in STAFF_SAMPLE:
        conn.execute("INSERT INTO staff (staff_name) VALUES (?)", (n,))


class TestFactSchema(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.conn = db.get_connection(os.path.join(self.tmp, "test.sqlite3"))
        db.init_db(self.conn)

    def tearDown(self):
        self.conn.close()

    def test_expected_tables_16(self):
        self.assertEqual(len(db.EXPECTED_TABLES), 16)
        tables = set(db.list_tables(self.conn))
        for t in db.EXPECTED_TABLES:
            self.assertIn(t, tables, f"필수 테이블 누락: {t}")
        self.assertIn("rulebook_facts", tables)
        self.assertIn("rulebook_fact_edits", tables)

    def test_init_is_idempotent(self):
        db.init_db(self.conn)
        db.init_db(self.conn)

    def test_review_status_defaults_to_unchecked(self):
        self.conn.execute(
            "INSERT INTO rulebook_facts (fact_kind, item_name, source_version) "
            "VALUES ('공통','x','V4.2')")
        r = self.conn.execute("SELECT review_status FROM rulebook_facts").fetchone()
        self.assertEqual(r["review_status"], "미확인")

    def test_review_status_limited_to_three(self):
        # D3: 미확인/확인함/보류 — 그 밖의 값은 창고가 거부해야 한다
        self.conn.execute(
            "INSERT INTO rulebook_facts (fact_kind, item_name, source_version) "
            "VALUES ('공통','y','V4.2')")
        for ok in ("확인함", "보류", "미확인"):
            self.conn.execute("UPDATE rulebook_facts SET review_status=? WHERE item_name='y'",
                              (ok,))
        with self.assertRaises(sqlite3.IntegrityError):
            self.conn.execute("UPDATE rulebook_facts SET review_status='대충확인' "
                              "WHERE item_name='y'")

    def test_item_name_unique_per_kind(self):
        self.conn.execute("INSERT INTO rulebook_facts (fact_kind, item_name, source_version) "
                          "VALUES ('공통','사회복지','V4.2')")
        # 종류가 다르면 같은 이름 허용(공통 카테고리명 = 개별 상품명이 겹칠 수 있다)
        self.conn.execute("INSERT INTO rulebook_facts (fact_kind, item_name, source_version) "
                          "VALUES ('개별','사회복지','V4.2')")
        with self.assertRaises(sqlite3.IntegrityError):
            self.conn.execute("INSERT INTO rulebook_facts (fact_kind, item_name, source_version) "
                              "VALUES ('공통','사회복지','V4.2')")


class TestFactLoadFromRulebook(unittest.TestCase):
    """실제 룰북 엑셀 → 임시 DB (실제 창고 파일은 건드리지 않는다)."""

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.tmp, "test.sqlite3")
        cls.xlsx = load_rulebook.find_rulebook_path()
        cls.detail1 = {}
        cls.counts1 = load_rulebook.run(cls.xlsx, cls.db_path, detail=cls.detail1)

    def setUp(self):
        self.conn = db.get_connection(self.db_path)

    def tearDown(self):
        self.conn.close()

    def test_fact_counts(self):
        self.assertEqual(self.counts1["facts"], 51)
        n = self.conn.execute(
            "SELECT fact_kind, COUNT(*) c FROM rulebook_facts GROUP BY fact_kind").fetchall()
        got = {r["fact_kind"]: r["c"] for r in n}
        self.assertEqual(got, {"공통": 16, "개별": 35})

    def test_all_start_unchecked(self):
        n = self.conn.execute(
            "SELECT COUNT(*) c FROM rulebook_facts WHERE review_status='미확인'").fetchone()["c"]
        self.assertEqual(n, 51)

    def test_item_names_present(self):
        # 식별키(항목명)가 빈 항목이 있으면 화면에서 부를 이름이 없다
        n = self.conn.execute(
            "SELECT COUNT(*) c FROM rulebook_facts "
            "WHERE item_name IS NULL OR TRIM(item_name)=''").fetchone()["c"]
        self.assertEqual(n, 0)

    def test_existing_loads_unbroken(self):
        # 회귀: 카테고리·금지어·개인정보 패턴 적재는 그대로여야 한다
        self.assertEqual(self.counts1["categories"], 15)
        self.assertEqual(self.counts1["banned_words"], 14)
        self.assertGreaterEqual(self.counts1["pii_patterns"], 6)

    def test_faq_cell_not_split(self):
        # 한 칸에 여러 개가 든 셀(FAQ TOP3)은 쪼개지 않고 그대로 보관
        rows = self.conn.execute(
            "SELECT faq_top3 FROM rulebook_facts "
            "WHERE fact_kind='공통' AND faq_top3 IS NOT NULL").fetchall()
        self.assertTrue(rows, "공통 팩트에 FAQ 칸이 하나도 없다")


class TestReloadPreservesReview(unittest.TestCase):
    """★ 가장 중요 — 다시 적재해도 검수 결과·화면 수정값이 살아 있어야 한다."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.db_path = os.path.join(self.tmp, "test.sqlite3")
        self.xlsx = load_rulebook.find_rulebook_path()
        self.counts1 = load_rulebook.run(self.xlsx, self.db_path)
        self.conn = db.get_connection(self.db_path)

    def tearDown(self):
        self.conn.close()

    def test_reload_keeps_count_status_and_edited_value(self):
        # 사람이 화면에서 한 일을 흉내: 한 항목을 '확인함'으로 + 값 수정
        fid = self.conn.execute(
            "SELECT fact_id FROM rulebook_facts WHERE fact_kind='공통' "
            "ORDER BY fact_id LIMIT 1").fetchone()["fact_id"]
        fixed = "2020.1.1 이전 120시간 / 이후 160시간"
        self.conn.execute(
            "UPDATE rulebook_facts SET review_status='확인함', "
            "reviewed_at=datetime('now','localtime'), faq_top3=? WHERE fact_id=?",
            (fixed, fid))
        self.conn.commit()
        self.conn.close()

        counts2 = load_rulebook.run(self.xlsx, self.db_path)  # 다시 적재
        self.conn = db.get_connection(self.db_path)

        self.assertEqual(counts2["facts"], 51)          # 중복 0
        self.assertEqual(self.counts1["facts"], counts2["facts"])
        row = self.conn.execute(
            "SELECT review_status, faq_top3 FROM rulebook_facts WHERE fact_id=?",
            (fid,)).fetchone()
        self.assertEqual(row["review_status"], "확인함")  # 도장 보존
        self.assertEqual(row["faq_top3"], fixed)          # 화면에서 고친 값 보존

    def test_reload_does_not_touch_posts(self):
        # 데이터 손실 가드: 적재가 기존 글 건수를 건드리면 안 된다
        self.conn.execute("INSERT INTO posts (title) VALUES ('기존 글')")
        self.conn.commit()
        before = self.conn.execute("SELECT COUNT(*) c FROM posts").fetchone()["c"]
        self.conn.close()
        load_rulebook.run(self.xlsx, self.db_path)
        self.conn = db.get_connection(self.db_path)
        after = self.conn.execute("SELECT COUNT(*) c FROM posts").fetchone()["c"]
        self.assertEqual(before, after)


class TestFactLoadRules(unittest.TestCase):
    """새 항목 추가 / 엑셀 변경 감지 — 임시 엑셀(메모리)로."""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.conn = db.get_connection(os.path.join(self.tmp, "test.sqlite3"))
        db.init_db(self.conn)
        load_rulebook.load_pii_patterns(self.conn)  # 마스킹 재료(정규식)
        _insert_staff(self.conn)                    # 마스킹 재료(이름 목록) — 실제 경로대로

    def tearDown(self):
        self.conn.close()

    def test_first_load_inserts_all_unchecked(self):
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        self.assertEqual((d["total"], d["inserted"], d["existing"]), (2, 2, 0))
        self.assertEqual(d["skipped_no_name"], 0)  # ▼ 안내 줄은 세지 않는다
        n = self.conn.execute("SELECT COUNT(*) c FROM rulebook_facts "
                              "WHERE review_status='미확인'").fetchone()["c"]
        self.assertEqual(n, 2)

    def test_new_row_added_existing_untouched(self):
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        self.conn.execute("UPDATE rulebook_facts SET review_status='확인함' "
                          "WHERE item_name='테스트상품'")
        new_row = (2, "개별", "테스트카테고리", "테스트상품2", "새 팩트",
                   "", "", "", "보통", "")
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1, new_row]))
        self.assertEqual((d["total"], d["inserted"], d["existing"]), (3, 1, 2))
        row = self.conn.execute("SELECT review_status FROM rulebook_facts "
                                "WHERE item_name='테스트상품'").fetchone()
        self.assertEqual(row["review_status"], "확인함")  # 기존 항목은 안 건드림
        new = self.conn.execute("SELECT review_status, core_fact FROM rulebook_facts "
                                "WHERE item_name='테스트상품2'").fetchone()
        self.assertEqual(new["review_status"], "미확인")
        self.assertEqual(new["core_fact"], "새 팩트")

    def test_changed_excel_is_reported_not_overwritten(self):
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        changed_row = list(IND_ROW_1)
        changed_row[4] = "실습 160시간"  # 엑셀 쪽 핵심 팩트가 바뀜
        d = load_rulebook.load_facts(self.conn, _fake_workbook([tuple(changed_row)]))
        self.assertEqual(d["inserted"], 0)
        self.assertIn("테스트상품", d["changed"])  # 알리기만
        row = self.conn.execute("SELECT core_fact FROM rulebook_facts "
                                "WHERE item_name='테스트상품'").fetchone()
        self.assertEqual(row["core_fact"], "실습 120시간")  # ★ 덮어쓰지 않음

    def test_unchanged_excel_reports_no_change(self):
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        self.assertEqual(d["changed"], [])
        self.assertEqual(d["renamed"], [])
        self.assertEqual(d["duplicate_names"], [])

    def test_renamed_item_detected_not_duplicated(self):
        # 엑셀에서 항목 '이름'만 고친 경우: 내용이 같으므로 중복으로 넣지 않고 무엇→무엇인지 알린다
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        row = list(IND_ROW_1)
        row[3] = "테스트 상품"  # '테스트상품' → '테스트 상품'
        d = load_rulebook.load_facts(self.conn, _fake_workbook([tuple(row)]))
        self.assertEqual(d["renamed"], [("테스트상품", "테스트 상품")])
        self.assertEqual(d["inserted"], 0)
        self.assertEqual(d["total"], 2)  # 공통1 + 개별1 — 미확인 새 항목이 생기지 않는다
        names = [r["item_name"] for r in self.conn.execute(
            "SELECT item_name FROM rulebook_facts WHERE fact_kind='개별'")]
        self.assertEqual(names, ["테스트상품"])  # 추정이므로 자동으로 이름을 바꾸지도 않는다

    def test_renamed_common_item_detected_not_duplicated(self):
        # 공통 팩트는 항목명이 곧 '카테고리' 칸이라, 내용 지문에서 그 칸을 빼야 이름 바뀜이 보인다
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        row = list(COMMON_ROW_1)
        row[2] = "테스트 카테고리"  # '테스트카테고리' → '테스트 카테고리' (이름만 바뀜)
        d = load_rulebook.load_facts(
            self.conn, _fake_workbook([IND_ROW_1], common_rows=[tuple(row)]))
        self.assertEqual(d["renamed"], [("테스트카테고리", "테스트 카테고리")])
        self.assertEqual(d["inserted"], 0)
        self.assertEqual(d["total"], 2)  # 공통1 + 개별1 — 미확인 새 항목이 생기지 않는다
        names = [r["item_name"] for r in self.conn.execute(
            "SELECT item_name FROM rulebook_facts WHERE fact_kind='공통'")]
        self.assertEqual(names, ["테스트카테고리"])  # 자동으로 이름을 바꾸지도 않는다

    def test_renamed_and_changed_common_item_becomes_new(self):
        # 공통도 이름·내용이 함께 바뀌면 알아볼 수 없다 → 새 항목(내용 지문이 이름만 빼는지 확인)
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        row = list(COMMON_ROW_1)
        row[2], row[4] = "테스트 카테고리", "150학점"
        d = load_rulebook.load_facts(
            self.conn, _fake_workbook([IND_ROW_1], common_rows=[tuple(row)]))
        self.assertEqual(d["renamed"], [])
        self.assertEqual(d["inserted"], 1)

    def test_empty_new_item_is_inserted_not_called_rename(self):
        # 이름·구분·연결카테고리만 채운 껍데기 행 둘 — 내용이 같다고 뒤엣것을 막으면 영영 못 넣는다
        stub_a = (2, "개별", "테스트카테고리", "신규상품A", "", "", "", "", "", "")
        stub_b = (3, "개별", "테스트카테고리", "신규상품B", "", "", "", "", "", "")
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1, stub_a]))
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1, stub_a, stub_b]))
        self.assertEqual(d["renamed"], [])       # A의 이름 바뀜으로 오판하면 안 된다
        self.assertEqual(d["inserted"], 1)       # B는 새 항목으로 들어간다
        names = [r["item_name"] for r in self.conn.execute(
            "SELECT item_name FROM rulebook_facts WHERE fact_kind='개별' ORDER BY item_name")]
        self.assertEqual(names, ["신규상품A", "신규상품B", "테스트상품"])

    def test_renamed_and_changed_becomes_new_item(self):
        # 이름도 내용도 바뀌면 알아볼 방법이 없다 → 새 항목으로 들어간다(정상 동작)
        load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1]))
        row = list(IND_ROW_1)
        row[3], row[4] = "테스트 상품", "실습 160시간"
        d = load_rulebook.load_facts(self.conn, _fake_workbook([tuple(row)]))
        self.assertEqual(d["renamed"], [])
        self.assertEqual(d["inserted"], 1)

    def test_duplicate_name_in_excel_reported(self):
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1, IND_ROW_1]))
        self.assertEqual(d["duplicate_names"], ["테스트상품"])
        self.assertEqual(d["total"], 2)  # 공통1 + 개별1

    def test_row_without_name_is_counted_not_dropped_silently(self):
        nameless = (9, "개별", "테스트카테고리", None, "이름 없는 팩트",
                    "", "", "", "", "")
        d = load_rulebook.load_facts(self.conn, _fake_workbook([IND_ROW_1, nameless]))
        self.assertEqual(d["skipped_no_name"], 1)
        self.assertEqual(d["inserted"], 2)  # 이름 있는 줄은 정상 적재

    def test_staff_name_in_fact_is_masked(self):
        # 불변 1 + reviewer 지적: 적재가 이름 목록 마스킹 경로를 실제로 밟는지 확인
        row = (4, "개별", "테스트카테고리", "테스트상품4", "담당 나철수 문의",
               "", "", "", "", "")
        d = load_rulebook.load_facts(self.conn, _fake_workbook([row]))
        got = self.conn.execute("SELECT core_fact FROM rulebook_facts "
                                "WHERE item_name='테스트상품4'").fetchone()
        self.assertNotIn("나철수", got["core_fact"])
        self.assertIn("[담당자]", got["core_fact"])
        self.assertTrue(any(h["type"] == "직원 실명/닉네임" for h in d["mask_hits"]))

    def test_masking_applied_to_facts(self):
        # 불변 1: 팩트도 프롬프트에 들어가므로 개인정보는 적재 시 가려져야 한다
        pii_row = (3, "개별", "테스트카테고리", "테스트상품3",
                   "문의는 010-1234-5678", "", "", "", "", "")
        d = load_rulebook.load_facts(self.conn, _fake_workbook([pii_row]))
        row = self.conn.execute("SELECT core_fact FROM rulebook_facts "
                                "WHERE item_name='테스트상품3'").fetchone()
        self.assertNotIn("010-1234-5678", row["core_fact"])
        self.assertIn("[전화번호]", row["core_fact"])
        self.assertTrue(d["mask_hits"])
        self.assertNotIn("original", d["mask_hits"][0])  # 가려진 원본값은 리포트에 담지 않음


class TestFactHeaderGuard(unittest.TestCase):
    """★ 조용한 0건 적재 차단 — 머리글이 바뀌면 그 종류가 통째로 빠지므로 멈춰야 한다."""

    def test_renamed_header_raises_instead_of_loading_zero(self):
        wb = _fake_workbook([IND_ROW_1], ind_name_header="상품명")  # '상품/키워드명'이 바뀜
        with self.assertRaisesRegex(ValueError, "개별"):
            load_rulebook.parse_fact_rows(wb)

    def test_empty_section_raises(self):
        with self.assertRaisesRegex(ValueError, "개별"):
            load_rulebook.parse_fact_rows(_fake_workbook([]))

    def test_missing_common_section_raises(self):
        with self.assertRaisesRegex(ValueError, "공통"):
            load_rulebook.parse_fact_rows(_fake_workbook([IND_ROW_1], common_rows=[]))

    def test_normal_sheet_does_not_raise(self):
        recs, _ = load_rulebook.parse_fact_rows(_fake_workbook([IND_ROW_1]))
        self.assertEqual(len(recs), 2)


class TestRunRollsBackOnBrokenExcel(unittest.TestCase):
    """★ 데이터 손실 가드 — 팩트에서 멈추면 카테고리·금지어 적재도 통째로 되돌아가야 한다.

    run()은 맨 끝에서 한 번만 commit한다 → 중간에 멈추면 그때까지의 DELETE·INSERT가 다 취소된다.
    '엑셀이 깨지면 통째로 멈춘다'가 안전한 이유가 이것이라, 여기서 못으로 박아둔다.
    """

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.db_path = os.path.join(self.tmp, "test.sqlite3")

    def _counts(self):
        conn = db.get_connection(self.db_path)
        try:
            return tuple(conn.execute(f"SELECT COUNT(*) c FROM {t}").fetchone()["c"]
                         for t in ("rulebook_categories", "rulebook_banned_words",
                                   "rulebook_facts"))
        finally:
            conn.close()

    def test_broken_excel_leaves_nothing_behind(self):
        broken = _full_workbook_file(self.tmp, "broken.xlsx", [])  # 개별 팩트 0건
        with self.assertRaisesRegex(ValueError, "개별"):
            load_rulebook.run(broken, self.db_path)
        self.assertEqual(self._counts(), (0, 0, 0))  # 반쪽 적재가 남으면 안 된다

    def test_broken_excel_does_not_wipe_previous_load(self):
        good = _full_workbook_file(self.tmp, "good.xlsx", [IND_ROW_1])
        counts = load_rulebook.run(good, self.db_path)
        self.assertEqual((counts["categories"], counts["banned_words"], counts["facts"]),
                         (1, 1, 2))
        broken = _full_workbook_file(self.tmp, "broken.xlsx", [])
        with self.assertRaises(ValueError):
            load_rulebook.run(broken, self.db_path)
        self.assertEqual(self._counts(), (1, 1, 2))  # 먼저 넣어둔 것이 지워지면 안 된다


class TestMaskingDoesNotBreakFacts(unittest.TestCase):
    """마스킹이 제도 수치·정상 문구를 훼손하면 안 된다(팩트는 원고에 그대로 들어간다).

    ★ 드라이런 결과(실제 이름 목록 × 팩트 시트 전 칸 → 가려진 칸 0)를 회귀로 고정한다.
    staff에 '지원'·'보람'처럼 흔한 낱말과 겹치는 이름이 들어오면 여기서 먼저 터져야 한다.
    (다만 이 테스트의 이름 목록은 표본이다. 실제 직원 이름이 바뀌었을 때의 신호는 적재 화면의
     '개인정보로 보여 가린 곳' 보고다.)
    """

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.conn = db.get_connection(os.path.join(self.tmp, "test.sqlite3"))
        db.init_db(self.conn)
        load_rulebook.load_pii_patterns(self.conn)
        _insert_staff(self.conn)
        self.pats = masking.load_regex_patterns(self.conn)
        self.names = masking.load_staff_names(self.conn)  # ★ 이름 목록을 실제로 넘긴다

    def tearDown(self):
        self.conn.close()

    def test_fact_sentences_unchanged(self):
        for s in FACT_SENTENCES:
            out, hits = masking.mask_text(s, self.pats, self.names)
            self.assertEqual(out, s, f"마스킹이 팩트 문장을 훼손함: {s}")
            self.assertEqual(hits, [])

    def test_junk_staff_values_do_not_eat_facts(self):
        # staff에 섞인 이름 아닌 값('7'·'린')이 숫자·낱말을 갉아먹으면 안 된다
        for s in ("필요 학점 7학점", "린스 무료 제공", "7일 이내 신청"):
            out, _ = masking.mask_text(s, self.pats, self.names)
            self.assertEqual(out, s, f"이름 아닌 staff 값이 팩트를 훼손함: {s}")

    def test_real_rulebook_facts_survive_name_masking(self):
        # 드라이런 재현: 실제 팩트 시트 전 칸 × 이름 목록 → 이름으로 가려진 칸 0
        wb = openpyxl.load_workbook(load_rulebook.find_rulebook_path(),
                                    read_only=True, data_only=True)
        try:
            d = load_rulebook.load_facts(self.conn, wb)  # 임시 DB에만 쓴다
        finally:
            wb.close()
        name_hits = [h for h in d["mask_hits"] if h["type"] == "직원 실명/닉네임"]
        self.assertEqual(name_hits, [], f"이름 마스킹이 실제 팩트 칸을 건드렸다: {name_hits[:5]}")
        self.assertEqual(d["inserted"], d["total"])

    def test_pii_in_fact_still_masked(self):
        out, _ = masking.mask_text("담당 가상인쌤에게 문의", self.pats, self.names)
        self.assertNotIn("가상인", out)


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    unittest.main(verbosity=2)
