# -*- coding: utf-8 -*-
"""load_rulebook.py — 룰북 V4.2 적재 (CA-1 최소 + CA-6 팩트)

적재 범위: 카테고리(①시트) + 금지어(④시트) + 개인정보 패턴 + 팩트(②시트, CA-6).
개인정보 패턴은 룰북 엑셀에 없으므로 서비스기획서 v9 §8·§8-2 정의를 seed로 넣는다.
`③ 매칭 규칙`·`⑤ 키워드 목록`은 아직 범위 밖. `계정 정보` 탭은 어떤 형태로도 안 읽는다(불변 2).

멱등:
  - 카테고리·금지어·개인정보 패턴: 같은 source_version 행을 지우고 다시 넣는다.
  - ★ 팩트(CA-6): **추가 전용 — DELETE 없음.** 화면에서 검수·수정한 내용이 최종본이라(계획 D1),
    재적재는 **새 항목만 INSERT**하고 기존 항목은 손대지 않는다. 엑셀 쪽 내용이 달라진 항목은
    덮어쓰지 않고 지문(fingerprint) 비교로 감지해 **알리기만 한다**. 내용은 같은데 이름만 다른
    항목은 '이름이 바뀐 것 같은 항목'으로 알리고 **새로 넣지 않는다**(추정이라 자동 변경도 안 함).

사용: python src/load_rulebook.py [룰북.xlsx 경로]
"""
import glob
import hashlib
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import masking  # noqa: E402
from db import ROOT_DIR, get_connection, init_db  # noqa: E402

SOURCE_VERSION = "V4.2"
CATEGORY_SHEET = "① 상품 카테고리"
BANNED_SHEET = "④ 금지어 및 CTA"
FACT_SHEET = "② 팩트 룰북"

# 개인정보·직원 식별정보 패턴 seed — 근거: 서비스기획서 v9 §8(마스킹)·§8-2(자동삽입 금지)
# 카페 닉네임·직원 실명은 정규식으로 못 잡으므로 name_list 자리만 마련(CA-3에서 목록 채움).
PII_PATTERN_SEEDS = [
    ("전화번호(일반)", "regex",
     r"0\d{1,2}[-.\s)]?\d{3,4}[-.\s]?\d{4}",
     "[전화번호]", "휴대폰·지역번호 형식 전화번호", "서비스기획서 v9 §8"),
    ("전화번호(대표번호)", "regex",
     r"1[568]\d{2}[-.\s]?\d{4}",
     "[전화번호]", "15xx/16xx/18xx 대표번호", "서비스기획서 v9 §8"),
    ("오픈채팅 링크", "regex",
     r"(?:https?://)?open\.kakao\.com/[^\s)\]}<>가-힣]+",
     "[오픈채팅링크]", "카카오 오픈채팅 URL(괄호·한글 앞에서 종료)", "서비스기획서 v9 §8"),
    ("직원 호칭(쌤/멘토/팀장)", "regex",
     r"[가-힣]{1,4}\s?(?:쌤|멘토|팀장)",
     "[담당자]", "OO쌤·OO멘토·OO팀장 등 내부 호칭", "서비스기획서 v9 §8-2"),
    ("직원 실명", "name_list",
     None,
     "[담당자]", "staff 테이블 확보 후 이름 목록으로 탐지(CA-3)", "서비스기획서 v9 §8"),
    ("카페 닉네임/활동명", "name_list",
     None,
     "[닉네임]", "계정별 활동명 목록 확보 후 탐지(CA-3)", "서비스기획서 v9 §8"),
]


def find_rulebook_path() -> str:
    """저장소 루트에서 룰북 엑셀을 찾는다."""
    hits = glob.glob(os.path.join(ROOT_DIR, "*룰북*.xlsx"))
    if not hits:
        raise FileNotFoundError("저장소 루트에서 '*룰북*.xlsx' 파일을 찾지 못했습니다.")
    return hits[0]


def load_categories(conn, wb) -> int:
    """①시트: No.가 숫자인 행만 카테고리로 적재."""
    ws = wb[CATEGORY_SHEET]
    conn.execute("DELETE FROM rulebook_categories WHERE source_version = ?", (SOURCE_VERSION,))
    count = 0
    for row in ws.iter_rows(values_only=True):
        no = row[0]
        if not isinstance(no, (int, float)):
            continue  # 제목·헤더 행 스킵
        top_cat, name, examples, link, kw_count, freq, active = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7])
        if not name:
            continue
        conn.execute(
            """INSERT INTO rulebook_categories
               (no, top_category, category_name, keyword_examples, credit_bank_link,
                unique_keyword_count, total_post_frequency, active, source_version)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (int(no), _s(top_cat), _s(name), _s(examples), _s(link),
             _i(kw_count), _i(freq),
             1 if _s(active) in (None, "Y", "y", "1") else 0,
             SOURCE_VERSION))
        count += 1
    return count


def load_banned_words(conn, wb) -> int:
    """④시트: '금지어 목록' 구간의 No.가 숫자인 행만 적재. 'CTA 문구 풀' 이후는 범위 밖."""
    ws = wb[BANNED_SHEET]
    conn.execute("DELETE FROM rulebook_banned_words WHERE source_version = ?", (SOURCE_VERSION,))
    count = 0
    in_banned_section = False
    for row in ws.iter_rows(values_only=True):
        first = _s(row[0])
        if first == "금지어 목록":
            in_banned_section = True
            continue
        if first == "CTA 문구 풀":
            break  # CTA는 CA-1 범위 밖
        if not in_banned_section:
            continue
        no = row[0]
        if not isinstance(no, (int, float)):
            continue  # 헤더 행(No. | 금지어 | ...) 스킵
        word, reason, replacement = _s(row[1]), _s(row[2]), _s(row[3])
        if not word:
            continue
        conn.execute(
            """INSERT INTO rulebook_banned_words (no, word, reason, replacement, source_version)
               VALUES (?, ?, ?, ?, ?)""",
            (int(no), word, reason, replacement, SOURCE_VERSION))
        count += 1
    return count


def load_pii_patterns(conn) -> int:
    """개인정보 패턴 seed 적재(v9 §8 기준)."""
    conn.execute("DELETE FROM rulebook_pii_patterns")
    for name, ptype, pattern, replacement, desc, source in PII_PATTERN_SEEDS:
        conn.execute(
            """INSERT INTO rulebook_pii_patterns
               (name, pattern_type, pattern, replacement, description, source)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (name, ptype, pattern, replacement, desc, source))
    return len(PII_PATTERN_SEEDS)


# ---------------------------------------------------------------------------
# ② 팩트 룰북 (CA-6)
# ---------------------------------------------------------------------------
# 엑셀 헤더명 → DB 칸. 헤더가 바뀌면 여기만 고치면 된다(비교는 공백 제거 후).
COMMON_HEADERS = {
    "응시/취득 요건": "requirement",
    "필요 학점": "credits",
    "예상 소요 기간": "duration",
    "기간 단축 방법": "shortcut",
    "자주 묻는 질문 TOP3": "faq_top3",
    "주의사항 / 흔한 오해": "cautions",
    "주의메모 (시점/예외)": "caution_memo",
    "구분": "division",
    "카테고리": "category",
}
INDIVIDUAL_HEADERS = {
    "핵심 팩트": "core_fact",
    "학력별 경로 요약": "path_by_education",
    "글 작성 시 강조포인트": "emphasis",
    "주의메모": "caution_memo",
    "사용 우선순위": "use_priority",
    "비고": "remarks",
    "구분": "division",
    "연결 카테고리": "category",
}
# 종류별: (헤더맵, 항목명 열 헤더) — 항목명이 식별키(공통=카테고리명 / 개별=상품·키워드명)
FACT_SECTIONS = {
    "공통": (COMMON_HEADERS, "카테고리"),
    "개별": (INDIVIDUAL_HEADERS, "상품/키워드명"),
}
# rulebook_facts의 내용 칸(식별·검수 컬럼 제외) — INSERT·지문 계산에 함께 쓴다
FACT_FIELDS = ("division", "category", "requirement", "credits", "duration", "shortcut",
               "faq_top3", "cautions", "caution_memo", "core_fact", "path_by_education",
               "emphasis", "use_priority", "remarks")


def _norm(s):
    """헤더 비교용 — 공백 전부 제거('주의사항 / 흔한 오해' == '주의사항/흔한오해')."""
    return re.sub(r"\s+", "", s or "")


def _sha(*parts):
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


def _name_field(kind):
    """그 종류의 '항목명'이 들어앉은 내용 칸(없으면 None).

    공통은 항목명이 곧 '카테고리' 칸이라(FACT_SECTIONS) 내용 지문에서 이 칸을 빼야 이름 바뀜을
    알아볼 수 있다. 개별은 항목명이 별도 칸(상품/키워드명)이라 뺄 것이 없다.
    """
    headers, name_header = FACT_SECTIONS[kind]
    return {_norm(k): v for k, v in headers.items()}.get(_norm(name_header))


def _fingerprint(kind, item_name, values):
    """엑셀 원본 내용의 지문. 엑셀 No.는 넣지 않는다(재번호는 내용 변경이 아니다).

    ★ 저장되는 지문의 정의(종류 + 항목명 + 내용칸 전부)는 바꾸지 않는다 — 바꾸면 이미 적재된
    항목이 전부 '엑셀이 바뀐 것 같음'으로 오인된다. 아래 내용 지문과 역할이 다르다.
    """
    return _sha(kind, item_name or "", *[values.get(f) or "" for f in FACT_FIELDS])


def _content_fingerprint(kind, values):
    """항목 '이름을 뺀' 내용만의 지문 — 이름만 바뀐 항목을 알아보는 데 쓴다(저장하지 않는다)."""
    skip = _name_field(kind)
    return _sha(*[values.get(f) or "" for f in FACT_FIELDS if f != skip])


# 이름·구분·연결카테고리만 채운 '껍데기' 행 판정용 — 이 칸들은 내용으로 치지 않는다
_STUB_FIELDS = ("division", "category")


def _has_content(kind, values):
    """이름·구분·연결카테고리 말고 실제 내용이 한 칸이라도 있나."""
    skip = set(_STUB_FIELDS) | {_name_field(kind)}
    return any(values.get(f) for f in FACT_FIELDS if f not in skip)


def _content_index(conn):
    """창고에 이미 있는 항목: (종류, 내용지문) → 항목명.

    내용이 사실상 빈 항목(이름·구분·연결카테고리만 있는 껍데기)은 서로 지문이 같아 엉뚱하게
    묶이므로 넣지 않는다 — 넣으면 내용 없는 진짜 새 항목이 앞엣것의 '이름만 바뀐 것'으로 오판돼
    영영 안 들어간다(강제로 넣을 방법이 없다).
    화면에서 사람이 값을 고친 항목은 엑셀과 내용이 달라져 여기 안 잡힌다(이름 바뀜 감지 실패
    → 새 항목으로 들어감). 추정을 넓히기보다 조용한 오판을 피하는 쪽을 택했다.
    """
    idx = {}
    sql = f"SELECT fact_kind, item_name, {', '.join(FACT_FIELDS)} FROM rulebook_facts"
    for r in conn.execute(sql):
        kind, values = r["fact_kind"], {f: r[f] for f in FACT_FIELDS}
        if kind in FACT_SECTIONS and _has_content(kind, values):
            idx.setdefault((kind, _content_fingerprint(kind, values)), r["item_name"])
    return idx


def parse_fact_rows(wb):
    """②시트 → [{'kind','excel_no','item_name', 내용칸...}] (마스킹 전).

    SECTION A/B 표시줄로 종류를 가르고, 그 아래 첫 'No.' 줄을 헤더로 잡는다.
    '▼ 아래부터 개별 팩트를 추가해주세요 ▼' 같은 안내 줄은 항목명이 없어 자연히 걸러진다
    (그 아래로 계속 늘어나는 행은 개별 팩트로 계속 읽힌다).

    ★ 종류(공통/개별) 하나라도 0건이면 예외를 던진다 — 엑셀 머리글이 바뀌면 그 종류가 통째로
    빠지는데, 그걸 "0건 적재"로 조용히 넘기면 팩트가 사라진 줄도 모른다.
    """
    ws = wb[FACT_SHEET]
    kind = None
    header_map = None      # 열 번호 → DB 칸 이름
    name_col = None        # 항목명 열 번호
    out, skipped_no_name = [], 0
    name_col_found = {k: False for k in FACT_SECTIONS}
    for row in ws.iter_rows(values_only=True):
        cells = [_s(c) for c in row]
        joined = " ".join(c for c in cells if c).upper()
        if "SECTION A" in joined:
            kind, header_map, name_col = "공통", None, None
            continue
        if "SECTION B" in joined:
            kind, header_map, name_col = "개별", None, None
            continue
        if kind is None:
            continue
        if joined.startswith("▼"):
            continue  # '▼ 아래부터 개별 팩트를 추가해주세요 ▼' 안내 줄
        headers, name_header = FACT_SECTIONS[kind]
        if header_map is None:  # 아직 헤더 줄을 못 만남
            if cells and _norm(cells[0]).upper() in ("NO.", "NO"):
                header_map, name_col = {}, None
                wanted = {_norm(k): v for k, v in headers.items()}
                for i, c in enumerate(cells):
                    key = _norm(c)
                    if key in wanted:
                        header_map[i] = wanted[key]
                    if key == _norm(name_header):
                        name_col = i
                name_col_found[kind] = name_col is not None
            continue
        if name_col is None or name_col >= len(cells):
            continue
        item_name = cells[name_col]
        if not item_name:
            if any(cells):  # 내용은 있는데 항목명이 빈 줄 — 조용히 버리지 않고 세어서 알린다
                skipped_no_name += 1
            continue
        rec = {"kind": kind, "item_name": item_name, "excel_no": _i(row[0])}
        for i, field in header_map.items():
            if i < len(cells) and cells[i]:
                rec[field] = cells[i]
        out.append(rec)

    for k, (_, name_header) in FACT_SECTIONS.items():
        if not name_col_found[k]:
            raise ValueError(
                f"'{FACT_SHEET}' 시트에서 {k} 팩트의 '{name_header}' 머리글을 찾지 못했습니다. "
                f"엑셀 머리글 이름이 바뀌었는지 확인해주세요 — 그대로 두면 {k} 팩트가 통째로 "
                f"빠집니다.")
        if not any(r["kind"] == k for r in out):
            raise ValueError(
                f"'{FACT_SHEET}' 시트에서 {k} 팩트를 한 건도 읽지 못했습니다. "
                f"'{name_header}' 칸이 채워져 있는지 확인해주세요 — 0건을 조용히 넘기지 않으려고 "
                f"멈춥니다.")
    return out, skipped_no_name


def load_facts(conn, wb) -> dict:
    """②시트 → rulebook_facts. ★ 추가 전용(DELETE·UPDATE 없음 — 검수 결과 보존).

    반환: {"total","inserted","existing","changed"(엑셀이 바뀐 것 같은 항목명),
           "renamed"[(옛이름, 새이름)], "duplicate_names","skipped_no_name","mask_hits"}
    """
    # 불변 1: 팩트는 나중에 AI 프롬프트에 그대로 들어간다 → 적재 시 마스킹 통과.
    regex_pats = masking.load_regex_patterns(conn)
    staff_names = masking.load_staff_names(conn)
    content_index = _content_index(conn)  # 이름만 바뀐 항목 감지용(적재 전 창고 기준)

    recs, skipped_no_name = parse_fact_rows(wb)
    inserted, existing, dups, mask_hits, seen = 0, 0, [], [], set()
    changed, renamed = [], []
    for rec in recs:
        kind = rec["kind"]
        # 가려진 원본값(hits[i]['original'])은 리포트에 담지 않는다 — 어느 항목·어느 칸인지만.
        item_name, hits = masking.mask_text(rec["item_name"], regex_pats, staff_names)
        mask_hits += [{"item": item_name, "field": "item_name", "type": h["type"]} for h in hits]
        values = {}
        for field in FACT_FIELDS:
            values[field], hits = masking.mask_text(rec.get(field), regex_pats, staff_names)
            mask_hits += [{"item": item_name, "field": field, "type": h["type"]} for h in hits]

        key = (kind, item_name)
        if key in seen:  # 엑셀 안에 같은 이름이 두 번 — 뒤엣것은 넣지 않고 알린다
            dups.append(item_name)
            continue
        seen.add(key)

        fp = _fingerprint(kind, item_name, values)
        row = conn.execute(
            "SELECT fact_id, source_fingerprint FROM rulebook_facts "
            "WHERE fact_kind=? AND item_name=?", (kind, item_name)).fetchone()
        if row is not None:
            existing += 1
            if row["source_fingerprint"] != fp:
                changed.append(item_name)  # ★ 덮어쓰지 않는다 — 건수만 알리고 사람이 화면에서 처리
            continue
        old_name = content_index.get((kind, _content_fingerprint(kind, values)))
        if old_name is not None:
            # 내용은 그대로인데 이름만 다르다 = 엑셀에서 항목명을 고친 것으로 보인다.
            # 추정일 뿐이라 ①새로 넣지 않고(검수한 옛 항목과 미확인 새 항목이 둘 다 남는 것 방지)
            # ②기존 이름을 자동으로 바꾸지도 않는다. 무엇이 무엇으로 보이는지만 알린다.
            renamed.append((old_name, item_name))
            continue
        cols = ["fact_kind", "excel_no", "item_name", "source_fingerprint", "source_version"]
        vals = [kind, rec.get("excel_no"), item_name, fp, SOURCE_VERSION]
        cols += list(FACT_FIELDS)
        vals += [values[f] for f in FACT_FIELDS]
        conn.execute(
            f"INSERT INTO rulebook_facts ({', '.join(cols)}) "
            f"VALUES ({', '.join('?' * len(cols))})", vals)
        inserted += 1  # review_status는 스키마 기본값 '미확인'

    total = conn.execute("SELECT COUNT(*) c FROM rulebook_facts").fetchone()["c"]
    return {"total": total, "inserted": inserted, "existing": existing, "changed": changed,
            "renamed": renamed, "duplicate_names": dups, "skipped_no_name": skipped_no_name,
            "mask_hits": mask_hits}


def _s(v):
    """셀 값 → 공백 정리 문자열(빈 값은 None)."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _i(v):
    return int(v) if isinstance(v, (int, float)) else None


def run(xlsx_path: str = None, db_path: str = None, detail: dict = None) -> dict:
    """룰북 적재 실행. 적재 건수 dict 반환(재실행해도 같은 숫자 = 멱등).

    detail dict를 넘기면 팩트 적재 상세(새로 넣은 건수·엑셀이 바뀐 것 같은 항목·마스킹 히트)를
    담아준다 — main()의 보고용. 반환 counts는 '지금 창고에 있는 건수'라 재실행해도 동일하다.
    """
    xlsx_path = xlsx_path or find_rulebook_path()
    conn = get_connection(db_path) if db_path else get_connection()
    init_db(conn)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        counts = {
            "categories": load_categories(conn, wb),
            "banned_words": load_banned_words(conn, wb),
            "pii_patterns": load_pii_patterns(conn),
        }
        fact_detail = load_facts(conn, wb)  # 개인정보 패턴 적재 뒤에 — 마스킹 재료가 필요
        counts["facts"] = fact_detail["total"]
        if detail is not None:
            detail.update(fact_detail)
        conn.commit()
    finally:
        wb.close()
        conn.close()
    return counts


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    xlsx = sys.argv[1] if len(sys.argv) > 1 else find_rulebook_path()
    print(f"룰북 파일: {os.path.basename(xlsx)}")
    detail = {}
    counts = run(xlsx, detail=detail)
    print(f"적재 완료 (source_version={SOURCE_VERSION}, 재실행 안전):")
    print(f"  - 카테고리      : {counts['categories']}건")
    print(f"  - 금지어        : {counts['banned_words']}건")
    print(f"  - 개인정보 패턴 : {counts['pii_patterns']}건")
    print(f"  - 팩트          : {counts['facts']}건 "
          f"(이번에 새로 넣은 것 {detail['inserted']}건 / 이미 있던 것 {detail['existing']}건)")
    if detail["changed"]:
        print(f"    * 엑셀이 바뀐 것 같은 항목 {len(detail['changed'])}건 — "
              f"창고 값은 그대로 뒀습니다(화면에서 확인하세요): "
              f"{', '.join(detail['changed'][:5])}"
              f"{' …' if len(detail['changed']) > 5 else ''}")
    if detail["renamed"]:
        print(f"    * 이름이 바뀐 것 같은 항목 {len(detail['renamed'])}건 — 내용이 같아서 "
              f"새로 넣지 않았습니다(창고 이름은 그대로, 화면에서 확인하세요):")
        for old, new in detail["renamed"]:
            print(f"      - '{old}' → '{new}'")
    if detail["duplicate_names"]:
        print(f"    * 엑셀에 이름이 겹치는 항목 {len(detail['duplicate_names'])}건 — "
              f"뒤엣것은 넣지 않았습니다: {', '.join(detail['duplicate_names'])}")
    if detail["skipped_no_name"]:
        print(f"    * 이름 칸이 빈 줄 {detail['skipped_no_name']}건 — 넣지 않았습니다"
              f"(공통은 '카테고리', 개별은 '상품/키워드명'을 채워주세요)")
    if detail["mask_hits"]:
        print(f"    * 개인정보로 보여 가린 곳 {len(detail['mask_hits'])}곳:")
        for h in detail["mask_hits"][:10]:
            print(f"      - {h['item']} / {h['field']} ({h['type']})")
    else:
        print("    * 개인정보로 가린 곳: 없음")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
